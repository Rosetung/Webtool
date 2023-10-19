from ast import Not
from pickle import POP
from statistics import mean
from turtle import ycor
import streamlit as st
import numpy as np
import pandas as pd
import plotly.express as px
import seaborn as sns
from scipy import stats
from scipy.stats import pearsonr
from sklearn import linear_model, metrics
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import csv
import openpyxl
from openpyxl import Workbook
import sys
from qstat import qstat
import plotly
import plotly.io as pio


st.set_page_config(layout="wide")
st.title('Make plotting flow cytometry data easier than GraphPad!')
st.write('To use this app, prepare an Excel file with the .xls or .xlsx file extension.')

# File uploader
#file = st.file_uploader('Upload your file:', type = ["xlsx", "csv", "xls"])

file_format = st.radio('Select file format:', ('csv', 'excel'), key='file_format')
file = st.file_uploader('Upload your file:')

if file is not None:
    #wb = openpyxl.load_workbook(file)
    #sheet_selector = st.selectbox('Select which sheet to open:', wb.sheetnames)
    if file_format == 'csv':
        flow_df = pd.read_csv(file) 
    else:
        wb = openpyxl.load_workbook(file)
        sheet_selector = st.selectbox('Select which sheet to open:', wb.sheetnames)
        flow_df = pd.read_excel(file, sheet_selector)
else:
    st.stop()

st.markdown('**Flow data**')
st.write(flow_df)

tab1, tab2, tab3 = st.tabs(["Bar graph", "Correlation analysis", "Statistical calculations"])

with tab1:
    # Select your groups
    st.subheader('Selection of categorical variables')

    non_num_cols = flow_df.select_dtypes(include=object).columns
    group_name = st.selectbox("Select categorical variable for analysis:", non_num_cols)
    remaining_non_num_cols = list(non_num_cols)
    remaining_non_num_cols.remove(group_name)
    remaining_non_num_cols.append("None")
    subgroup_name = st.selectbox("Select sub-category for analysis:", remaining_non_num_cols)
    grouplist = flow_df[group_name]
    if subgroup_name != "None":
        subgrouplist = flow_df[subgroup_name]

    unique_values = flow_df.value_counts(group_name if subgroup_name == 'None' else [group_name, subgroup_name]).reset_index()
    unique_values = unique_values.rename(columns={ 0: 'count'})

    col1, col2 = st.columns([1, 2])
    with col1: 
        st.markdown('***Unique counts for each condition:***')
        st.write(unique_values)

    with col2:
        color_val = subgroup_name if subgroup_name != 'None' else None
        groups_fig = px.bar(unique_values, x=group_name, y='count',color=color_val)
        st.plotly_chart(groups_fig)

    # Select your numerical parameters and your groups for display/plotting
    st.subheader('Numerical parameters for analysis')

    unique_groups = grouplist.unique()
    selected_groups = st.multiselect('Select groups to compare:', unique_groups, unique_groups)
    if subgroup_name != "None":
        unique_subgroups = subgrouplist.unique()
        selected_subgroups = st.multiselect('Select sub-groups to compare:', unique_subgroups, unique_subgroups)

    if subgroup_name != "None":
        selected_data_df = flow_df[flow_df[group_name].isin(selected_groups)][flow_df[subgroup_name].isin(selected_subgroups)]
    else:
        selected_data_df = flow_df[flow_df[group_name].isin(selected_groups)]
    st.write(selected_data_df)

    numerical_parameters = list(flow_df.select_dtypes(include='number').columns)
    selected_parameters = st.multiselect('Select parameters to plot:', numerical_parameters, numerical_parameters)

    def getChart(selected_parameter):
        # Draw the bar chart
        fig_selected_parameter = plt.figure(figsize=(24, 18))
        hue_val = subgroup_name if subgroup_name != 'None' else None
        b=sns.barplot(
            data=selected_data_df, 
            x=group_name, 
            y=selected_parameter, 
            hue=hue_val,
            errorbar="se",
            errcolor="0",
            errwidth=6, 
            linewidth=8,
            edgecolor="0", 
            capsize=.2
        )
        plt.ylabel(selected_parameter, fontsize=40)
        plt.tick_params(width=3, labelsize=40)
        plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0, fontsize=20)
        b.set(xlabel=None)
        b.spines['left'].set_linewidth(5)
        b.spines['bottom'].set_linewidth(5)
        b.spines['right'].set_visible(False)
        b.spines['top'].set_visible(False)
        
        
        

        # Draw the stripplot
        hue_val1 = subgroup_name if subgroup_name != 'None' else group_name
        c=sns.stripplot(
            data=selected_data_df, 
            x=group_name, 
            y=selected_parameter, 
            hue=hue_val1, 
            jitter=True,
            dodge=subgroup_name != 'None', 
            edgecolor="black", 
            linewidth=8,
            s=30,
            legend=False 
        )
        c.set(xlabel=None)

        return fig_selected_parameter

    cols = st.columns([1, 1])
    for idx, selected_parameter in enumerate(selected_parameters):
        with cols[idx % 2]:
            fig_selected_paramter = getChart(selected_parameter)
            st.pyplot(fig_selected_paramter)

    def downloadFigures():
        for selected_parameter in selected_parameters:
            getChart(selected_parameter)    # savefig will save the last created chart
            plt.savefig(selected_parameter + '.eps', format='eps')
    
    st.button('Download Figures', on_click=downloadFigures)

with tab2:
    # Draw scatterplot
    st.header("Scatterplot with correlation")

    selected_x_var = st.selectbox('What do you want the x variable to be?', options=numerical_parameters)
    selected_y_var = st.selectbox('What about the y?', options=numerical_parameters)
    fig = px.scatter(flow_df, x = flow_df[selected_x_var], y = flow_df[selected_y_var])
    st.plotly_chart(fig)
    def downloadchart():
        fn='scatterplot.svg'
        ff=pio.write_image(fig, fn)
    st.button('Download Scatter plot', on_click=downloadchart)


    # Correlation test
    def calc_corr(selected_x_var, selected_y_var):
        corr, p_val = stats.pearsonr(selected_x_var, selected_y_var)
        return corr, p_val

    st.subheader("Pearson correlation coefficient")

    x = flow_df[selected_x_var].to_numpy()
    y = flow_df[selected_y_var].to_numpy()

    if file is not None:
        dataset_used = flow_df
        correlation, corr_p_val = calc_corr(x, y)
        st.write('Pearson correlation = %.3f' % correlation)
        st.write('p-value = %.3f' % corr_p_val)

    def calc_corr(selected_x_var, selected_y_var):
        corr, p_val = stats.spearmanr(selected_x_var, selected_y_var)
        return corr, p_val

    st.subheader("Spearman's rank correlation coefficient")

    x = flow_df[selected_x_var].to_numpy()
    y = flow_df[selected_y_var].to_numpy()

    if file is not None:
        dataset_used = flow_df
        correlation, corr_p_val = calc_corr(x, y)
        st.write('Spearman correlation = %.3f' % correlation)
        st.write('p-value = %.3f' % corr_p_val)


    st.header("Heatmap of correlation matrix")
    selected_multiple_parameters = st.multiselect('Select parameters to plot. If nothing is selected, all parameters will be shown.', options=numerical_parameters)

    if len(selected_multiple_parameters) != 0:
        selected_data_df1 = flow_df[selected_multiple_parameters]
        st.write(selected_data_df1)

        corr = selected_data_df1.corr()
        with sns.axes_style("white"):
            f, ax = plt.subplots(figsize=(10, 7))
            ax = sns.heatmap(corr, cmap="vlag", center=0, square=True, linewidths=.5, annot=True, cbar_kws={"shrink": .5})
        st.write(f)
        def downloadplots():
            ht = f.get_figure()
            fa='heatmap.svg'
            aa=ht.savefig(fa)
        st.button('Download Heatmap', on_click=downloadplots)
                
        

        st.header("Pairwise plot")
        st.caption("Parameters will be the same as selected previously.")
        selected_color = st.selectbox('Colored by:', ["None"] + non_num_cols.to_list())
        if selected_color != "None":
            g = sns.pairplot(flow_df, vars=selected_data_df1, dropna = True, hue=selected_color, diag_kind="kde")
            g.map_lower(sns.regplot)
            st.pyplot(g)
        else:
            g = sns.pairplot(flow_df, vars=selected_data_df1, dropna = True, diag_kind="kde")
            g.map_lower(sns.regplot)
            st.pyplot(g)
        def downloadplot():
            hg = g.fig
            fpp='pairplot.svg'
            bb=hg.savefig(fpp)
        st.button('Download Pairwise plot', on_click=downloadplot)

    else:
        corr = flow_df.corr()
        with sns.axes_style("white"):
            f, ax = plt.subplots(figsize=(10, 7))
            ax = sns.heatmap(corr, cmap="vlag", center=0, square=True, linewidths=.5, annot=True, cbar_kws={"shrink": .5})
        st.write(f)
        def downloadplots():
            ht = f.get_figure()
            fa='heatmap.svg'
            aa=ht.savefig(fa)
        st.button('Download Heatmap', on_click=downloadplots)
        
        st.header("Pairwise plot")
        selected_color = st.selectbox('Colored by', ["None"] + non_num_cols.to_list())
        if selected_color != "None":
            g = sns.pairplot(flow_df, dropna = True, hue=selected_color, diag_kind="kde")
            g.map_lower(sns.regplot)
            st.pyplot(g)
        else:
            g = sns.pairplot(flow_df, dropna = True, diag_kind="kde")
            g.map_lower(sns.regplot)
            st.pyplot(g)
        def downloadplot():
            hg = g.fig
            fpp='pairplot.svg'
            bb=hg.savefig(fpp)
        st.button('Download Pairwise plot', on_click=downloadplot)


with tab3:
     # Statistics here we go!
    import math
    import scipy 
    from scipy.stats.morestats import shapiro
    import pandas as pd
    import numpy as np
    from plotnine import ggplot, aes, geom_line, coord_fixed, geom_segment, geom_col, geom_boxplot, coord_flip
    import pingouin as pg
    import statistics as stats
    from qstat import qstat


    # title of the app
    t_choice = st.radio("Choose statistical test",["One Sample Data","Paired T-test","Unpaired T-test", "One-way ANOVA", "Two-way ANOVA"])
    if t_choice == "One Sample Data":
        
        st.subheader("One Sample Data - Descriptive analysis")
        c1,c2,c3 = st.columns((1,1,1))

        with c1: 
            df = flow_df.copy()
            df = df.dropna(axis=1, how="all")
            # st.dataframe(df.assign(hack='').set_index('hack')) 
            global numeric_columns
            global non_numeric_columns
            numeric_columns = df.select_dtypes(include=np.number).columns.tolist()
            non_numeric_columns = list(df.select_dtypes(['object']).columns)
            
            quant = st.selectbox('Numerical variable', options=numeric_columns)
            cat = st.selectbox('Main Categorical variable', options=non_numeric_columns)


        with c2:    
            if cat != None:
                allcat = list(df[cat].unique())
                cat1 = st.selectbox('Choose group to analyze',options=allcat) 
                sdf = df[[quant,cat]]
                fsdf = sdf[sdf[cat]==cat1]
                # st.markdown(f"Numerical variable: {quant}")
                # st.markdown(f"Categorical variable: {cat}")   
                # st.markdown(f"Group: {cat1}")  
                st.dataframe(fsdf.describe())
                data_one_sample = df[quant]
                
            if cat == None:
                fsdf = df[quant]
                st.markdown(f"Numerical variable: {quant}")
                st.markdown(f"Categorical variable: {cat}")  
                qstat = pd.DataFrame(fsdf.describe())
                st.write(qstat)
            fsdf=pd.DataFrame(fsdf)
            
            
        with c3:
            st.markdown("Normality test")
            p = pg.qqplot(fsdf[quant], dist='norm')
            st.pyplot(ggplot.draw(p))
            ddd = shapiro(fsdf[quant])[1]
            st.write("Shapiro p-Value: " + str(ddd))
            if ddd<0.05:
                st.write("Data does not have normal distribution. Please transform data in original file and reupload for analysis for parametric tests to be done.")
            else:
                st.write("Data has normal distribution.")
        
        st.markdown('''---''')
        st.subheader("One sample data - inferential analysis (significance test)")
        import scipy.stats
        import pandas as pd
        t_stats, p_value = scipy.stats.ttest_1samp(a=fsdf[quant], popmean=5000)
        st.write("T-statistics: " + str(t_stats))
        st.write("P-value: " + str(p_value))
    
    if t_choice == "Paired T-test":
        st.subheader("Paired Sample Data - Descriptive analysis")
        c1,c2,c3 = st.columns((1,2,1))
        with c1:
            df = flow_df.copy()
            df = df.dropna(axis=1, how="all")
             
            #global numeric_columns
            #global non_numeric_columns
            numeric_columns = list(df.select_dtypes(['float', 'int']).columns)
            non_numeric_columns = list(df.select_dtypes(['object']).columns)
               
            qb = st.selectbox('Before treatment', options=numeric_columns)
            qa = st.selectbox('After treatment', options=numeric_columns,index = 1)
            cat = st.selectbox('Categorical variable', options=non_numeric_columns)
        with c1:
            df['After-Before'] = df[qa]-df[qb]
            quant = 'After-Before'
            # st.dataframe(df.assign(hack='').set_index('hack')) 
        with c2:
            if cat != None:
                allcat = list(df[cat].unique())
                cat1 = st.selectbox('Choose group to analyze',options=allcat) 
                sdf = df[[quant,cat]]
                fsdf = sdf[sdf[cat]==cat1]
                # st.markdown(f"Quantity: {quant}")
                # st.markdown(f"Category: {cat}")   
                # st.markdown(f"Variable: {cat1}")  
                st.dataframe(fsdf.describe())
            if cat == None:
                fsdf = df[quant]
                # st.markdown(f"Quantity: {quant}")
                # st.markdown(f"Category: {cat}")  
                st.write(pd.DataFrame(fsdf.describe()))
            fsdf=pd.DataFrame(fsdf)
        with c3:
            st.markdown("Normality test")
            p = pg.qqplot(fsdf[quant], dist='norm')
            st.pyplot(ggplot.draw(p))
            ddd = shapiro(fsdf[quant])[1]
            st.write("Shapiro p-Value: " + str(ddd))
            if ddd<0.05:
                st.write("Data does not have normal distribution. Please transform data in original file and reupload for analysis for parametric tests to be done.")
            else:
                st.write("Data has normal distribution.")
        
        
        st.markdown('''---''')
        st.subheader("Paired sample data - inferential analysis (significance test)")
        d1,d2,d3 = st.columns((1,2,3))
        with d1:
            nh = float(st.text_input("Null Hypothesis:",0))
            alpha = float(st.text_input("Alpha:",0.05))
            tail_choice = st.radio("",["Left Tail","Two Tails","Right Tail"])
        with d2:
            n = len(fsdf)
            df = n-1
            xbar = stats.mean(fsdf[quant])
            s = stats.stdev(fsdf[quant])
            sem = s/math.sqrt(n)
            ts = (xbar - nh)/sem
            x = np.arange(-5,5,.1)
            ty = scipy.stats.t.pdf(x,df)
            tdf = pd.DataFrame({"x":x,"ty":ty})
            tplot = ggplot(tdf) + coord_fixed(ratio = 4)
            if tail_choice == "Left Tail":
                pvalue = scipy.stats.t.cdf(ts,df)
                cv = scipy.stats.t.ppf(alpha,df)
                tdf["Left"] = np.where(tdf["x"]<=ts,tdf["ty"],0)
                tplot = tplot + geom_col(aes(x=x,y="Left"), fill = "steelblue", size = .1, alpha = .4)
                cl = 1 - 2*alpha
            if tail_choice == "Two Tails":
                rts = abs(ts)
                lts = -rts
                pvalue = 2*scipy.stats.t.cdf(lts,df)
                cv = scipy.stats.t.ppf(alpha/2,df)
                cv = abs(cv)
                tdf["Center"] = np.where(np.logical_or(tdf["x"]>=rts,tdf["x"]<=lts),tdf["ty"],0)
                tplot = tplot + geom_col(aes(x=x,y="Center"), fill = "steelblue", size = .1, alpha = .4)
                cl = 1-alpha
            if tail_choice == "Right Tail":
                pvalue = 1-scipy.stats.t.cdf(ts,df)
                cv = scipy.stats.t.ppf(1-alpha,df)
                tdf["Right"] = np.where(tdf["x"]>=ts,tdf["ty"],0)
                tplot = tplot + geom_col(aes(x=x,y="Right"), fill = "steelblue", size = .1, alpha = .4)
                cl=1-2*alpha
            me = cv*sem
            data = pd.DataFrame({"n":n,"df":df,"x-bar":xbar,"s":s,"sem":sem,"CV t*":cv,"ME":me,"t-Score":ts,"p-Value":pvalue},index = [0]).T 
            st.write(data) 
        with d3:
            tplot = tplot + geom_segment(aes(x = ts, y = 0, xend = ts, yend = scipy.stats.t.pdf(ts,df)),color="red")
            tplot = tplot + geom_line(aes(x=x,y=ty)) + coord_fixed(ratio = 4)
            st.pyplot(ggplot.draw(tplot))
            lower = xbar - abs(me)
            upper = xbar + abs(me) 
            st.write(str(100*cl) + "'%' confidence interval is (" + str(lower) +", "+str(upper)+")") 
            
    if t_choice == "Unpaired T-test":
        st.markdown('*The results below will be based on the selection in the Bar graph tab*')
        if len(selected_groups)==2:
            if subgroup_name == 'None':
                tab1, tab2, tab3 = st.tabs(["Descriptive statistics", "Normality test", "Inferential statistics"])
                with tab1:
                    st.subheader("Descriptive statistics")
                    st.markdown("Descriptive statistics for chosen numerical parameters")
                    dff=selected_data_df[selected_parameters]
                    dfff = dff.describe(include='all')
                    st.write(dfff)

                with tab2:
                    st.subheader("Normality test")
                    cols = st.columns([1, 1])
                    with cols[idx % 1]:
                        numbers=range(len(selected_parameters))
                        for jj in numbers:
                            dsf=selected_data_df[selected_parameters[jj]]
                            st.write(selected_parameters[jj])
                            p = pg.qqplot(dsf, dist='norm')
                            st.pyplot(ggplot.draw(p))
                            shap1 = scipy.stats.shapiro(dsf)[1]
                            st.write("Shapiro p-Value: " + str(shap1))
                            if shap1<0.05:
                                st.write("Data does not have normal distribution. Please transform data in original file and reupload for analysis for parametric tests to be done.")
                            else:
                                st.write("Data has normal distribution.")
            
                with tab3:
                    st.subheader("Inferential statistics")
                    #gp1=selected_data_df[group_name==selected_groups[0]]
                    gp1 = selected_data_df[selected_data_df[group_name]==selected_groups[0]]
                    gp2 = selected_data_df[selected_data_df[group_name]==selected_groups[1]]
                    #st.write(gp1)
                    for jj in numbers:
                        t_stats2, p_val2 = scipy.stats.ttest_ind(gp1[selected_parameters[jj]], gp2[selected_parameters[jj]], equal_var=True)
                        st.markdown(selected_parameters[jj])
                        st.write('T-statistics: ' + str(t_stats2))
                        st.write('P-value: ' + str(p_val2))
                        st.write('')
            if subgroup_name != 'None':
                if len(selected_subgroups) == 1:
                    tab1, tab2, tab3 = st.tabs(["Descriptive statistics", "Normality test", "Inferential statistics"])
                    with tab1:
                        st.subheader("Descriptive statistics")
                        st.markdown("Descriptive statistics for chosen numerical parameters")
                        dff=selected_data_df[selected_parameters]
                        dfff = dff.describe(include='all')
                        st.write(dfff)

                    with tab2:
                        st.subheader("Normality test")
                        cols = st.columns([1, 1])
                        with cols[idx % 1]:
                            numbers=range(len(selected_parameters))
                            for jj in numbers:
                                dsf=selected_data_df[selected_parameters[jj]]
                                st.write(selected_parameters[jj])
                                p = pg.qqplot(dsf, dist='norm')
                                st.pyplot(ggplot.draw(p))
                                shap1 = scipy.stats.shapiro(dsf)[1]
                                st.write("Shapiro p-Value: " + str(shap1))
                                if shap1<0.05:
                                    st.write("Data does not have normal distribution. Please transform data in original file and reupload for analysis for parametric tests to be done.")
                                else:
                                    st.write("Data has normal distribution.")
                
                    with tab3:
                        st.subheader("Inferential statistics")
                        #gp1=selected_data_df[group_name==selected_groups[0]]
                        gp1 = selected_data_df[selected_data_df[group_name]==selected_groups[0]]
                        gp2 = selected_data_df[selected_data_df[group_name]==selected_groups[1]]
                        #st.write(gp1)
                        for jj in numbers:
                            t_stats2, p_val2 = scipy.stats.ttest_ind(gp1[selected_parameters[jj]], gp2[selected_parameters[jj]], equal_var=True)
                            st.markdown(selected_parameters[jj])
                            st.write('T-statistics: ' + str(t_stats2))
                            st.write('P-value: ' + str(p_val2))
                            st.write('')


            
        elif subgroup_name == 'None' and len(selected_groups)>2:
            st.write('Perform One-way ANOVA')  
        elif subgroup_name != 'None' and len(selected_subgroups) == 1:
            st.write('Perform One-way ANOVA')
        elif subgroup_name != 'None' and len(selected_groups)>1 and len(selected_subgroups) > 1:
            st.write('Perform Two-way ANOVA')

        else:
            st.write('Not enough groups to compare')

    # ANOVA
    import statsmodels
    import statsmodels.api as sm
    import scikit_posthocs as sp
    from statsmodels.stats.weightstats import ttest_ind
    from statsmodels.stats.multicomp import pairwise_tukeyhsd
    import pingouin as pg
    

    # One-way ANOVA
    
    if t_choice == "One-way ANOVA":
        st.markdown('*The results below will be based on the selection in the Bar graph tab*')
        if subgroup_name == 'None':
            if len(selected_groups) > 2:
                numbers=range(len(selected_parameters))
                agree = st.checkbox('Perform posthoc Tukey multiple comparison test')
                for jj in numbers:
                    if agree:
                        aov = pg.anova(dv=selected_parameters[jj], between=group_name, data= selected_data_df, detailed=True)
                        hs_res=pairwise_tukeyhsd(selected_data_df[selected_parameters[jj]], selected_data_df[group_name])
                        pf = pd.DataFrame(data = hs_res._results_table.data[1: ], columns = hs_res._results_table.data[0])
                        st.write(selected_parameters[jj]) 
                        st.write('- One-way ANOVA', aov)
                        st.write('- Posthoc Tukey multiple comparison test',pf)
                    else:
                        aov = pg.anova(dv=selected_parameters[jj], between=group_name, data= selected_data_df, detailed=True)
                        st.write(selected_parameters[jj]) 
                        st.write('- One-way ANOVA', aov)
    
            if len(selected_groups) < 3:
                st.write('Not enough groups to compare. Requires more than 3 groups')
        if subgroup_name != 'None':
            if len(selected_subgroups) == 1:
                numbers=range(len(selected_parameters))
                agree = st.checkbox('Perform posthoc Tukey multiple comparison test')
                for jj in numbers:
                    if agree:
                        aov = pg.anova(dv=selected_parameters[jj], between=group_name, data= selected_data_df, detailed=True)
                        hs_res=pairwise_tukeyhsd(selected_data_df[selected_parameters[jj]], selected_data_df[group_name])
                        pf = pd.DataFrame(data = hs_res._results_table.data[1: ], columns = hs_res._results_table.data[0])
                        st.write(selected_parameters[jj]) 
                        st.write('- One-way ANOVA', aov)
                        st.write('- Posthoc Tukey multiple comparison test',pf)
                    else:
                        aov = pg.anova(dv=selected_parameters[jj], between=group_name, data= selected_data_df, detailed=True)
                        st.write(selected_parameters[jj]) 
                        st.write('- One-way ANOVA', aov)
        else:
            st.write('Please perform Two-way ANOVA')

        
    # Two-way ANOVA
    
    
    if t_choice == "Two-way ANOVA":
        st.markdown('*The results below will be based on the selection in the Bar graph tab*')
        if subgroup_name != 'None':
            
            numbers=range(len(selected_parameters))
            agree = st.checkbox('Perform posthoc Tukey multiple comparison test')
            for jj in numbers:
                if agree:
                    aov = pg.anova(dv=selected_parameters[jj], between=[group_name, subgroup_name], data= selected_data_df, detailed=True)
                    selected_data_df['combination'] = selected_data_df[group_name]+ " / " + selected_data_df[subgroup_name]
                    res = pairwise_tukeyhsd(selected_data_df[selected_parameters[jj]], selected_data_df['combination'], alpha=0.05)
                    pp = pd.DataFrame(data = res._results_table.data[1: ], columns = res._results_table.data[0])
            
                    st.write(selected_parameters[jj])
                    st.write('- Two-way ANOVA', aov)
                    st.write('- Posthoc Tukey multiple comparison test',pp)
                else:
                    aov = pg.anova(dv=selected_parameters[jj], between=[group_name, subgroup_name], data= selected_data_df, detailed=True)
                    st.write(selected_parameters[jj])
                    st.write('- Two-way ANOVA', aov)
                

        else:
            if subgroup_name == 'None':
                st.write('Please select sub-category to compare.')
                if len(selected_groups) > 2:
                    st.write('If not, please perfom One-way ANOVA.')
                if len(selected_groups) < 3:
                    st.write('If not, please perfom one sample, paired sample or two sample data tests.')
            
        
